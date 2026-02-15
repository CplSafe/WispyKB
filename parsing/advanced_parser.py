"""
高级文档解析模块
参考：RAGFlow, Dify 的文档解析实现

功能：
1. PDF表格提取
2. 图片OCR
3. 多模态文档解析
4. 表格结构化输出
"""

import logging
import re
from typing import List, Dict, Any, Optional, Union
from dataclasses import dataclass
from io import BytesIO
import json

logger = logging.getLogger(__name__)


@dataclass
class TableData:
    """表格数据"""
    headers: List[str]
    rows: List[List[str]]
    markdown: str
    metadata: Dict[str, Any]


@dataclass
class ImageData:
    """图片数据"""
    content: str  # OCR文本
    format: str
    metadata: Dict[str, Any]


class AdvancedDocumentParser:
    """
    高级文档解析器

    支持：
    - PDF表格提取
    - 图片OCR
    - Word表格
    - Excel/CSV
    - Markdown表格
    """

    def __init__(self, ocr_enabled: bool = True):
        """
        初始化解析器

        Args:
            ocr_enabled: 是否启用OCR
        """
        self.ocr_enabled = ocr_enabled

        # 尝试导入可选依赖
        self.pdfplumber_available = False
        self.pymupdf_available = False
        self.pil_available = False
        self.pytesseract_available = False

        try:
            import pdfplumber
            self.pdfplumber_available = True
        except ImportError:
            pass

        try:
            import fitz  # PyMuPDF
            self.pymupdf_available = True
        except ImportError:
            pass

        try:
            from PIL import Image
            self.pil_available = True
        except ImportError:
            pass

        try:
            import pytesseract
            self.pytesseract_available = True
        except ImportError:
            pass

    def parse_pdf_tables(self, file_path: str) -> List[TableData]:
        """
        解析PDF表格

        Args:
            file_path: PDF文件路径

        Returns:
            表格数据列表
        """
        tables = []

        if self.pdfplumber_available:
            import pdfplumber
            try:
                with pdfplumber.open(file_path) as pdf:
                    for page_num, page in enumerate(pdf.pages):
                        page_tables = page.extract_tables()
                        for table_num, table in enumerate(page_tables):
                            # 转换为TableData
                            table_data = self._convert_pdf_table(table, page_num, table_num)
                            if table_data:
                                tables.append(table_data)
                logger.info(f"Extracted {len(tables)} tables from PDF using pdfplumber")
            except Exception as e:
                logger.error(f"PDF table extraction error: {e}")

        return tables

    def _convert_pdf_table(self, raw_table: List, page_num: int, table_num: int) -> Optional[TableData]:
        """转换PDF表格为TableData"""
        if not raw_table or len(raw_table) < 2:
            return None

        # 第一行作为表头
        headers = [str(cell) if cell else "" for cell in raw_table[0]]
        rows = []
        for row in raw_table[1:]:
            if row:
                rows.append([str(cell) if cell else "" for cell in row])

        # 生成Markdown表格
        markdown = self._table_to_markdown(headers, rows)

        return TableData(
            headers=headers,
            rows=rows,
            markdown=markdown,
            metadata={"page": page_num + 1, "table": table_num + 1}
        )

    def _table_to_markdown(self, headers: List[str], rows: List[List[str]]) -> str:
        """转换为Markdown表格格式"""
        lines = []

        # 表头
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

        # 数据行
        for row in rows:
            lines.append("| " + " | ".join(row) + " |")

        return "\n".join(lines)

    def parse_excel(self, file_path: str) -> List[TableData]:
        """
        解析Excel文件

        Args:
            file_path: Excel文件路径

        Returns:
            表格数据列表
        """
        tables = []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = []
                for row in ws.iter_rows(values_only=True):
                    if row and any(cell is not None for cell in row):
                        data.append([str(cell) if cell is not None else "" for cell in row])

                if len(data) >= 2:
                    headers = data[0]
                    rows = data[1:]
                    markdown = self._table_to_markdown(headers, rows)

                    tables.append(TableData(
                        headers=headers,
                        rows=rows,
                        markdown=markdown,
                        metadata={"sheet": sheet_name}
                    ))

            logger.info(f"Extracted {len(tables)} tables from Excel")
        except ImportError:
            logger.warning("openpyxl not available")
        except Exception as e:
            logger.error(f"Excel parsing error: {e}")

        return tables

    def parse_csv(self, file_path: str) -> Optional[TableData]:
        """
        解析CSV文件

        Args:
            file_path: CSV文件路径

        Returns:
            表格数据
        """
        import csv

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                data = list(reader)

            if len(data) >= 2:
                headers = data[0]
                rows = data[1:]
                markdown = self._table_to_markdown(headers, rows)

                return TableData(
                    headers=headers,
                    rows=rows,
                    markdown=markdown,
                    metadata={"source": "csv"}
                )
        except Exception as e:
            logger.error(f"CSV parsing error: {e}")

        return None

    def parse_markdown_tables(self, content: str) -> List[TableData]:
        """
        解析Markdown中的表格

        Args:
            content: Markdown内容

        Returns:
            表格数据列表
        """
        tables = []

        # 匹配Markdown表格
        table_pattern = re.compile(
            r'^\|(?P<headers>[^\n]+)\|\n\|(?P<separator>[\s\-:]+)\|\n(?P<rows>(?:\|[^\n]+\|\n?)+)',
            re.MULTILINE
        )

        for match in table_pattern.finditer(content):
            headers = [h.strip() for h in match.group('headers').split('|') if h.strip()]
            rows_text = match.group('rows').strip()
            rows = []

            for line in rows_text.split('\n'):
                if line.strip():
                    cells = [c.strip() for c in line.split('|') if c.strip()]
                    if cells:
                        rows.append(cells)

            if headers and rows:
                markdown = self._table_to_markdown(headers, rows)
                tables.append(TableData(
                    headers=headers,
                    rows=rows,
                    markdown=markdown,
                    metadata={"source": "markdown"}
                ))

        logger.info(f"Extracted {len(tables)} tables from Markdown")
        return tables

    def extract_text_from_image(self, image_data: bytes) -> Optional[str]:
        """
        从图片提取文字（OCR）

        Args:
            image_data: 图片二进制数据

        Returns:
            识别的文字
        """
        if not self.ocr_enabled:
            return None

        if self.pil_available and self.pytesseract_available:
            try:
                from PIL import Image
                import pytesseract

                image = Image.open(BytesIO(image_data))
                text = pytesseract.image_to_string(image, lang='chi_sim+eng')
                return text.strip()
            except Exception as e:
                logger.error(f"OCR error: {e}")

        return None

    def parse_document(self, file_path: str, file_type: str) -> Dict[str, Any]:
        """
        智能解析文档

        Args:
            file_path: 文件路径
            file_type: 文件类型

        Returns:
            解析结果 {text: str, tables: list, images: list}
        """
        result = {
            "text": "",
            "tables": [],
            "images": [],
            "metadata": {"file_type": file_type}
        }

        if file_type == "pdf":
            result = self._parse_pdf(file_path)
        elif file_type in ["xlsx", "xls"]:
            tables = self.parse_excel(file_path)
            result["tables"] = [t.markdown for t in tables]
            result["text"] = "\n\n".join([t.markdown for t in tables])
        elif file_type == "csv":
            table = self.parse_csv(file_path)
            if table:
                result["tables"] = [table.markdown]
                result["text"] = table.markdown
        elif file_type in ["md", "markdown"]:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            result["text"] = content
            result["tables"] = [t.markdown for t in self.parse_markdown_tables(content)]

        return result

    def _parse_pdf(self, file_path: str) -> Dict[str, Any]:
        """解析PDF文档"""
        result = {"text": "", "tables": [], "images": []}

        # 提取文本
        if self.pymupdf_available:
            import fitz
            try:
                doc = fitz.open(file_path)
                for page in doc:
                    result["text"] += page.get_text()
                doc.close()
            except Exception as e:
                logger.error(f"PDF text extraction error: {e}")

        # 提取表格
        tables = self.parse_pdf_tables(file_path)
        result["tables"] = [t.markdown for t in tables]

        return result


# 全局解析器实例
advanced_parser = AdvancedDocumentParser()


# 工具函数
def extract_tables_from_content(content: str) -> str:
    """
    从内容中提取并格式化表格

    Args:
        content: 包含表格的文本

    Returns:
        格式化后的表格文本
    """
    tables = advanced_parser.parse_markdown_tables(content)
    if tables:
        return "\n\n".join([f"表格{i+1}:\n{t.markdown}" for i, t in enumerate(tables)])
    return ""


def parse_file(file_path: str, file_type: str = None) -> Dict[str, Any]:
    """
    便捷函数：解析文件

    Args:
        file_path: 文件路径
        file_type: 文件类型（可选，自动检测）

    Returns:
        解析结果
    """
    if file_type is None:
        file_type = file_path.split('.')[-1].lower()

    return advanced_parser.parse_document(file_path, file_type)
